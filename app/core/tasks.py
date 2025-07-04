import os
from typing import Optional

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy import select, join

from app.core.celery_app import celery_app
from app.core.database import Transaction, Category


sync_engine = create_engine("postgresql://postgres:654321@db:5432/finance_db")

SyncSession = sessionmaker(bind=sync_engine)


@celery_app.task
def process_task_celery(user_id) -> Optional[str]:
    with SyncSession() as session:
        j = join(Transaction, Category, Transaction.id_category == Category.id)
        query = select(
            Transaction.id, 
            Category.name, 
            Transaction.type, 
            Transaction.amount, 
            Transaction.date,
            Transaction.description,
            ).select_from(j).where(Transaction.id_user == user_id)
        result = session.execute(query)
        result = result.mappings().all()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '№'
    ws['B1'] = 'name'
    ws['C1'] = 'amount'
    x = 1
    for row in result:
        if row['type'] == 'income':
            ws.append((x, row['name'], row['amount']))
        else:
            ws.append((x, row['name'], -(row['amount'])))
        x += 1
    ws.append(['', 'Total:', f'=ROUND(SUM(C2:C{x}), 2)'])
    folder_path = 'reports'
    file_name = f'report_user_id_{user_id}.xlsx'
    full_path = os.path.join(folder_path, file_name)
    wb.save(full_path)
    return f'Report created: {full_path}'

    